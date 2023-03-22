# 3322 输入动捕的总帧率 让传感器对齐动捕
# 输入动捕开启时间的千分秒
# 输入动捕结束时间的千分秒
# 读取传感器 第一列为千分秒
# 匹配最接近开始的第一个千分秒
# 匹配最接近结束的最后一个千分秒
#
# 设 x为动捕总帧率 y为两匹配间的帧率 t为开始匹配 T为结束匹配
# 每秒种应过滤的帧数 H
# 过滤算法 ： H =  (y-x)/T-t
# Testdata : t = 2023/01/09_14:48:02.120 T = 2023/01/09_14:48:57.120

from datetime import datetime
import openpyxl
from openpyxl import load_workbook

frame_optrick = int(input("请输入动捕的总帧率:"))
# 读取Excel表格中的数据
wb = load_workbook("./process_20230227/0227_001_R_Sensor.xlsx")
ws = wb["Sheet"]


def save_new_excel(x1, y1, x2, y2, ws):

    # 用于动捕与传感器数据正好相等的情况 重新写入新excel
    data = []
    for row in range(x1, x2 + 1):
        row_data = []
        for col in range(y1, y2 + 1):
            row_data.append(ws.cell(row=row, column=col).value)
        data.append(row_data)

    wk = openpyxl.Workbook()
    sheet = wk.active
    for i in range(len(data)):
        for j in range(len(data[i])):
            sheet.cell(row=x1 + i, column=y1 + j).value = data[i][j]
    wk.save("newfile.xlsx")
    wk.close


# 获取第一列第3行开始的千分秒序列
def add_line_save_excel(x1, y1, x2, y2, ws, number):
    # 用于每多少行添加一行数据
    # 添加数据与上一行一致
    data = []
    for row in range(x1, x2 + 1):
        row_data = []
        for col in range(y1, y2 + 1):
            row_data.append(ws.cell(row=row, column=col).value)
        data.append(row_data)

    new_data = []
    for i in range(len(data)):
        new_data.append(data[i])
        if (i + 1) % number == 0:
            new_data.append(data[i])
    wk = openpyxl.Workbook()
    sheet = wk.active
    for i in range(len(new_data)):
        for j in range(len(new_data[i])):
            sheet.cell(row=x1 + i, column=y1 + j).value = new_data[i][j]
    wk.save("newfile_add.xlsx")
    wk.close


def delete_line_save_excel(x1, y1, x2, y2, ws, number):
    # 用于每多少行添加一行数据
    # 添加数据与上一行一致
    data = []
    for row in range(x1, x2 + 1):
        row_data = []
        for col in range(y1, y2 + 1):
            row_data.append(ws.cell(row=row, column=col).value)
        data.append(row_data)

    new_data = []
    for i in range(len(data)):
        if (i + 1) % number != 0:
            new_data.append(data[i])
    wk = openpyxl.Workbook()
    sheet = wk.active
    for i in range(len(new_data)):
        for j in range(len(new_data[i])):
            sheet.cell(row=x1 + i, column=y1 + j).value = new_data[i][j]
    wk.save("./process_20230227/0227_001_R_Sensor_process.xlsx")
    wk.close


microseconds = []
for row in ws.iter_rows(min_row=3, min_col=1, max_col=1):
    microsecond = row[0].value
    microsecond = datetime.strptime(
        microsecond, "%Y/%m/%d_%H:%M:%S.%f"
    )  # 将字符串转化为千分秒类型准备进行计算
    microseconds.append(microsecond)
t = input("输入动捕开始的千分秒时间 例子：2023/01/09_14:48:02.120 :")
t = datetime.strptime(t, "%Y/%m/%d_%H:%M:%S.%f")
len_microseconds = len(microseconds)  # 获得列表长度
find_start_time = []
for mi in microseconds:
    find_start_time.append(abs(t - mi))
min_find_start_time = min(find_start_time)  # 获得最小差值
start_time_index = find_start_time.index(min_find_start_time)  # 获得最小值索引
print(start_time_index)

T = input("输入动捕结束的千分秒时间 例子：2023/01/09_14:48:57.120 :")
T = datetime.strptime(T, "%Y/%m/%d_%H:%M:%S.%f")

find_end_time = []
# print(microseconds)
for et in microseconds:
    find_end_time.append(abs(T - et))
min_find_end_time = min(find_end_time)
end_time_rever_index = find_end_time.index(min_find_end_time)  # 获得翻转后的结束index
print(end_time_rever_index)
Leg_frame = end_time_rever_index - start_time_index  # 获得传感器在这段时间内的帧率

if Leg_frame == frame_optrick:
    print("帧率已对齐，无需进行操作，准备将对齐后的传感器数据写入文件种")
    save_new_excel(3, 4, 3 + Leg_frame, 11, ws)
    print("写入完成")
if Leg_frame < frame_optrick:
    print("传感器少于动捕帧率，需要补充帧率")
    print("缺少帧率为:", frame_optrick - Leg_frame)
    add_frame = frame_optrick - Leg_frame
    add_location_number = int(
        Leg_frame / add_frame
    )  # 需要添加的总数 例如3500/5 = 700 即为每700帧添加一位
    print("add_number", add_location_number)
    add_line_save_excel(3, 4, 3 + Leg_frame, 11, ws, add_location_number)
    print("补充完成 已保存")
if Leg_frame > frame_optrick:
    print("传感器帧率多于动捕帧率，将进行删除")
    print("多出来的帧率为:", Leg_frame - frame_optrick)
    delete_frame = Leg_frame - frame_optrick
    delete_frame_number = int(Leg_frame / delete_frame)
    print("delete", delete_frame_number)
    delete_line_save_excel(3, 4, 3 + Leg_frame, 11, ws, delete_frame_number)
    print("删除完成")
